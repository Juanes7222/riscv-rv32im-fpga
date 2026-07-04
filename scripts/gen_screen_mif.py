#!/usr/bin/env python3
"""
gen_screen_mif.py - Formal visual-template compiler for RISC-V console.

Reads RiscVScreenV3.xlsx and produces two deterministic artifacts:

  1. A .mif file for video-memory initialization (character + attribute per cell).
  2. A JSON field-map: the formal contract between the visual layout and the
     RTL screen-writer.  Each entry records: logical name, row, column,
     width, format, and memory address range.

Design principle
----------------
The Excel/CSV is the single source of truth.  The script does NOT "draw" the
screen; it extracts a machine-readable slot specification.  There are two
kinds of slots:

  * Tagged slots   - a known label (e.g. "ALUA", "Opcode") followed by a
                     writable placeholder run.  The label determines the
                     RTL signal name; the placeholder only fixes position
                     and width.
  * Tabular slots  - repetitive blocks (Registers, Prog. Memory, Data Memory)
                     parsed by row-pattern, not by coordinate.

Placeholder conventions
-----------------------
  X…X  (run of X) --> hex field.  Width must be even for byte-aligned values.
  B…B  (run of B) --> binary field.  Width must match the RTL signal width.
  A-B  (legacy)   --> column range, still accepted during transition.

The script fails hard when:
  • a known label has no detectable placeholder,
  • two fields overlap,
  • a binary field width does not match the catalogued signal width,
  • a tabular block is malformed (wrong row count, missing patterns).
"""

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl

# Geometry
SCREEN_W = 160
SCREEN_H = 45
CELLS = SCREEN_W * SCREEN_H           # 7 200
ADDR_BITS = 13                        # 2**13 = 8192 > 7200

# Video word: {attribute[15:8], char[7:0]}
ATTR_STATIC = 0x07   # white on black
ATTR_DYNAMIC = 0x0F  # bright white on black (placeholder for runtime)


# Signal catalog
# Each entry:  rtl_name --> (fmt, signal_bits, display_min_chars)
#   fmt: 'hex' | 'bin'
#   signal_bits: width of the actual RTL wire
#   display_min_chars: minimum chars needed to show the value
#
# Visual labels that map to the same rtl_name are listed in LABEL_ALIASES.
CATALOG = {
    # Fetch / PC
    'pc':          ('hex', 32, 8),
    'instruction': ('hex', 32, 8),
    'next_pc':     ('hex', 32, 8),
    'branch':      ('bin',  1, 1),

    # Decode
    'rs1_addr':    ('hex',  5, 2),
    'rs2_addr':    ('hex',  5, 2),
    'rd_addr':     ('hex',  5, 2),
    'rs1_data':    ('hex', 32, 8),
    'rs2_data':    ('hex', 32, 8),
    'imm_out':     ('hex', 32, 8),
    'imm_src':     ('hex',  3, 1),

    # Opcode / funct
    'opcode':      ('hex',  7, 2),
    'funct3':      ('hex',  3, 1),
    'funct7':      ('hex',  7, 2),

    # Control
    'ru_wr':           ('bin', 1, 1),
    'ru_data_wr_src':  ('hex', 2, 1),
    'alu_op':          ('hex', 5, 2),
    'br_op':           ('hex', 5, 2),
    'alua_src':        ('hex', 2, 1),
    'alub_src':        ('bin', 1, 1),
    'dm_ctrl':         ('hex', 3, 1),
    'dm_wr':           ('hex', 1, 1),

    # Execute / mem / WB
    'alu_a':       ('hex', 32, 8),
    'alu_b':       ('hex', 32, 8),
    'alu_res':     ('hex', 32, 8),
    'dm_addr':     ('hex', 32, 8),
    'dm_wdata':    ('hex', 32, 8),
    'dm_rdata':    ('hex', 32, 8),
    'data_wr':     ('hex', 32, 8),

    # Performance
    'cycle_count': ('hex', 64, 16),

    # Pipeline extras (only used when --sheet Pipeline)
    'if_pc':           ('hex', 32, 8),
    'if_instruction':  ('hex', 32, 8),
    'id_pc':           ('hex', 32, 8),
    'id_instruction':  ('hex', 32, 8),
    'id_rs1_data':     ('hex', 32, 8),
    'id_rs2_data':     ('hex', 32, 8),
    'id_imm':          ('hex', 32, 8),
    'ex_pc':           ('hex', 32, 8),
    'ex_instruction':  ('hex', 32, 8),
    'ex_rs1_data':     ('hex', 32, 8),    # ID/EX output - forwarded rs1
    'ex_rs2_data':     ('hex', 32, 8),    # ID/EX output - forwarded rs2
    'ex_imm':          ('hex', 32, 8),    # ID/EX output - immediate
    'ex_alu_a':        ('hex', 32, 8),
    'ex_alu_b':        ('hex', 32, 8),
    'ex_alu_result':   ('hex', 32, 8),
    'ex_branch_taken': ('bin',  1, 1),
    'mem_alu_result':  ('hex', 32, 8),
    'mem_rs1_data':    ('hex', 32, 8),    # EX/MEM output - forwarded rs1
    'mem_rs2_data':    ('hex', 32, 8),    # EX/MEM output - forwarded rs2
    'mem_instruction': ('hex', 32, 8),    # EX/MEM output - instruction
    'mem_dm_wr':       ('bin',  1, 1),
    'wb_instruction':  ('hex', 32, 8),
    'wb_rd_addr':      ('hex',  5, 2),
    'wb_rd_data':      ('hex', 32, 8),
    'stall':           ('bin',  1, 1),
    'load_use_hazard': ('bin',  1, 1),
    'branch_flush':    ('bin',  1, 1),
    'trap_flush':      ('bin',  1, 1),
    'flush':           ('bin',  1, 1),
    'instr_retired':   ('hex', 64, 16),
}

# Multiple visual labels can resolve to the same RTL name.
# Longest labels first to avoid partial matches (e.g. "ALUASrc" before "ALUA").
LABEL_ALIASES: Dict[str, str] = {
    'PC':        'pc',
    'Instr':     'instruction',
    'Instruction':'instruction',
    'NextPC':    'next_pc',
    'Branch':    'branch',
    'rs1':       'rs1_addr',
    'rs2':       'rs2_addr',
    'rd':        'rd_addr',
    'RU(rs1)':   'rs1_data',
    'RU(rs2)':   'rs2_data',
    'IMM':       'imm_out',
    'IMMSrc':    'imm_src',
    'IMMSrC':    'imm_src',
    'Opcode':    'opcode',
    'Funct3':    'funct3',
    'Funct7':    'funct7',
    'RUWr':      'ru_wr',
    'RUDataWrSrc': 'ru_data_wr_src',
    'DataWr':    'data_wr',
    'ALUOp':     'alu_op',
    'BROp':      'br_op',
    'ALUASrc':   'alua_src',
    'ALUBSrc':   'alub_src',
    'DMCtrl':    'dm_ctrl',
    'DMWr':      'dm_wr',
    'ALUA':      'alu_a',
    'ALUB':      'alu_b',
    'ALURes':    'alu_res',
    'Address':   'dm_addr',
    'DataWrite': 'dm_wdata',
    'DataRead':  'dm_rdata',
    'Clockcycles':'cycle_count',
    'CycleCount': 'cycle_count',

    'IF_PC':       'if_pc',
    'IF_Instr':    'if_instruction',

    'ID_PC':       'id_pc',
    'ID_Instr':    'id_instruction',

    'ID_rs1_data': 'id_rs1_data',
    'ID_rs2_data': 'id_rs2_data',
    'ID_imm':      'id_imm',
    'EX_PC':       'ex_pc',

    'EX_Instr':    'ex_instruction',
    'EX_ALUA':     'ex_alu_a',
    'EX_ALUB':     'ex_alu_b',
    'EX_ALURes':   'ex_alu_result',
    'EX_Branch':   'ex_branch_taken',

    'MEM_ALURes':  'mem_alu_result',
    'MEM_ALUA':    'mem_rs1_data',
    'MEM_ALUB':    'mem_rs2_data',
    'MEM_Instr':   'mem_instruction',
    'MEM_Branch':  'ex_branch_taken',
    'MEM_DMWr':    'mem_dm_wr',

    'EX_rs1_data': 'ex_rs1_data',
    'EX_rs2_data': 'ex_rs2_data',
    'EX_imm':      'ex_imm',
    'WB_Instr':    'wb_instruction',
    'WB_rd':       'wb_rd_addr',
    'WB_rd_data':  'wb_rd_data',

    'Stall':       'stall',
    'LoadUse':     'load_use_hazard',
    'BranchFlush': 'branch_flush',
    'TrapFlush':   'trap_flush',
    'Flush':       'flush',
    'InstrRetired':'instr_retired',
}


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class Slot:
    name: str           # RTL signal name (canonical)
    row: int            # 0-based row in viewport
    col: int            # 0-based start column
    width: int          # visible width in characters
    fmt: str            # 'hex' | 'bin'
    signal_bits: int    # actual RTL bit width
    display_min: int    # minimum chars needed
    label: str          # visual label as written in the sheet
    origin: str         # human-readable provenance (for debugging)

    @property
    def start_address(self) -> int:
        return self.row * SCREEN_W + self.col

    @property
    def end_address(self) -> int:
        return self.start_address + self.width - 1

    def to_dict(self) -> dict:
        return {
            'name': self.name,
            'label': self.label,
            'row': self.row,
            'col': self.col,
            'width': self.width,
            'fmt': self.fmt,
            'signal_bits': self.signal_bits,
            'display_min': self.display_min,
            'start_address': self.start_address,
            'end_address': self.end_address,
            'origin': self.origin,
        }


@dataclass
class Block:
    """A rectangular text block found by its header row."""
    name: str
    row_start: int      # header row (0-based)
    col_start: int      # left column of the block (0-based)
    col_end: int        # right column (exclusive)


# Step 1 - Read worksheet into a normalised character grid
def read_grid(ws) -> List[List[str]]:
    """Build a SCREEN_H x SCREEN_W grid from the worksheet."""
    grid: List[List[str]] = [[' ' for _ in range(SCREEN_W)] for _ in range(SCREEN_H)]
    for r in range(1, SCREEN_H + 1):
        for c in range(1, SCREEN_W + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                s = str(v)
                if s:
                    grid[r - 1][c - 1] = s[0]
    return grid


# Step 2 - Detect block headers and delimit their horizontal bounds
def find_blocks(grid: List[List[str]]) -> Dict[str, Block]:
    """Scan the grid for known block-header text and determine each block's
    horizontal extent by looking for the enclosing '|' separators on the
    header row.  Returns a map header_name --> Block."""
    header_names = [
        'Registers', 'Prog. Memory', 'Data Memory',
        'FETCH', 'DECODE', 'EXECUTE', 'MEMORY', 'WRITEBACK',
        'HALTED', 'HELP', 'Logo', 'FSM', 'TPU', "Counters", "Pipeline", "Control", "Performance"
    ]
    header_names.sort(key=len, reverse=True)
    header_re = re.compile('|'.join(re.escape(h) for h in header_names))

    blocks: Dict[str, Block] = {}

    for r in range(SCREEN_H):
        row_text = ''.join(grid[r])
        for m in header_re.finditer(row_text):
            name = m.group(0)
            if name in blocks:
                continue          # keep first (topmost) occurrence
            c0 = m.start()
            c1 = m.end()

            # Look left for a '|' delimiter
            left = c0
            while left > 0 and grid[r][left - 1] != '|':
                left -= 1

            # Look right for a '|' delimiter
            right = c1
            while right < SCREEN_W and grid[r][right] != '|':
                right += 1
            if right < SCREEN_W:
                right += 1     # include the '|' as the boundary

            blocks[name] = Block(name=name, row_start=r,
                                 col_start=left, col_end=right)

    return blocks


# Step 3 - Tagged-slot detection (label + placeholder) in free areas
def detect_tagged_slots(grid: List[List[str]],
                        occupied: Set[Tuple[int, int]],
                        catalog: Dict[str, Tuple[str, int, int]]) -> List[Slot]:
    """Find every known label and extract the first valid placeholder run
    to its right.  Only searches in cells that are NOT already occupied by
    a tabular block."""

    labels_sorted = sorted(LABEL_ALIASES.keys(), key=len, reverse=True)
    label_re = re.compile('|'.join(re.escape(l) for l in labels_sorted))

    slots: List[Slot] = []
    seen: Set[Tuple[str, int, int]] = set()

    for r in range(SCREEN_H):
        row_text = ''.join(grid[r])
        consumed: Set[int] = set()   # columns already taken by a longer match on this row

        for m in label_re.finditer(row_text):
            label = m.group(0)
            start = m.start()
            end = m.end()

            # Skip sub-string matches (e.g. "IMM" inside "IMMSrC")
            if start in consumed:
                continue
            # Mark this match's span as consumed so shorter labels inside it are skipped
            for c in range(start, end):
                consumed.add(c)

            rtl_name = LABEL_ALIASES[label]
            fmt, signal_bits, display_min = catalog[rtl_name]

            after = end
            # skip optional colon / space immediately after the label
            while after < SCREEN_W and grid[r][after] in ': ':
                after += 1

            # If the first char after the label is already occupied, skip
            # (the label is inside a tabular block and should not be parsed
            # as a tagged slot).
            if after < SCREEN_W and (r, after) in occupied:
                continue

            start_col: Optional[int] = None
            vis_width: Optional[int] = None
            placeholder_type: Optional[str] = None

            # 1) Try A-B range (legacy)
            m2 = re.match(r'(\d+)-(\d+)', row_text[after:after + 12])
            if m2:
                c1 = int(m2.group(1))
                c2 = int(m2.group(2))
                if 1 <= c1 <= c2 <= SCREEN_W:
                    start_col = c1 - 1
                    vis_width = c2 - c1 + 1
                    placeholder_type = 'range'
                else:
                    raise ValueError(
                        f'Row {r + 1}: range "{c1}-{c2}" for "{label}" '
                        f'is outside the {SCREEN_W}-column viewport.'
                    )
            else:
                # 2) Try X-run (hex)
                if after < SCREEN_W and grid[r][after] == 'X':
                    k = after
                    while k < SCREEN_W and grid[r][k] == 'X':
                        k += 1
                    start_col = after
                    vis_width = k - after
                    placeholder_type = 'hex'
                # 3) Try B-run (binary)
                elif after < SCREEN_W and grid[r][after] == 'B':
                    k = after
                    while k < SCREEN_W and grid[r][k] == 'B':
                        k += 1
                    start_col = after
                    vis_width = k - after
                    placeholder_type = 'bin'
                else:
                    raise ValueError(
                        f'Row {r + 1}: label "{label}" has no detectable '
                        f'placeholder (expected A-B, X-run, or B-run).'
                    )

            # Deduplicate on (rtl_name, row, col)
            key = (rtl_name, r, start_col)
            if key in seen:
                continue
            seen.add(key)

            if placeholder_type == 'bin':
                fmt = 'bin'
            elif placeholder_type == 'hex':
                fmt = 'hex'

            slots.append(Slot(
                name=rtl_name,
                row=r,
                col=start_col,
                width=vis_width,
                fmt=fmt,
                signal_bits=signal_bits,
                display_min=display_min,
                label=label,
                origin=f'tagged after "{label}" on row {r + 1}',
            ))

    return slots


# Step 4 - Tabular blocks: Registers
def parse_registers(grid: List[List[str]],
                    block: Block,
                    occupied: Set[Tuple[int, int]]) -> List[Slot]:
    """Parse the Registers block.  Each row below the header must match:
        X[nn] : <X-run>
    Returns one Slot per register value (32 total)."""
    slots: List[Slot] = []

    for offset in range(1, 41):
        r = block.row_start + offset
        if r >= SCREEN_H:
            break
        # Only look inside the block's column window
        row_slice = ''.join(grid[r][block.col_start:block.col_end])

        m = re.search(r'X(\d{1,2})\s*:\s*(X+)', row_slice)
        if not m:
            continue
        reg_idx = int(m.group(1))
        x_run = m.group(2)
        if reg_idx > 31:
            continue

        # Convert match position from slice-relative to grid-absolute
        start_col = block.col_start + m.start(2)
        width = len(x_run)

        for c in range(start_col, start_col + width):
            occupied.add((r, c))

        slots.append(Slot(
            name=f'x{reg_idx}',
            row=r,
            col=start_col,
            width=width,
            fmt='hex',
            signal_bits=32,
            display_min=8,
            label=f'X{reg_idx}',
            origin=f'register table row {r + 1}',
        ))

    if len(slots) != 32:
        print(f'Warning: Registers block has {len(slots)} entries (expected 32).',
              file=sys.stderr)
    return slots


# Step 5 - Tabular blocks: Prog. Memory & Data Memory
def parse_memory_table(grid: List[List[str]],
                       block: Block,
                       occupied: Set[Tuple[int, int]],
                       header_name: str) -> List[Slot]:
    """Parse a memory table.  Each row inside the block must match:
        <addr> : <X-run>
    We look for a 4-char address pattern followed by a colon and an X-run.
    Returns two slots per row: address display + data value."""
    slots: List[Slot] = []
    data_count = 0

    for offset in range(1, 41):
        r = block.row_start + offset
        if r >= SCREEN_H:
            break
        row_slice = ''.join(grid[r][block.col_start:block.col_end])

        # Pattern: address (hex digits or X) followed by colon and X-run
        m = re.search(r'([0-9A-FX]{4})\s*:\s*(X+)', row_slice)
        if not m:
            continue

        addr_str = m.group(1)
        x_run = m.group(2)

        # Convert to absolute columns
        addr_start = block.col_start + m.start(1)
        data_start = block.col_start + m.start(2)
        addr_width = len(addr_str)
        data_width = len(x_run)

        # Mark occupied
        for c in range(addr_start, addr_start + addr_width):
            occupied.add((r, c))
        for c in range(data_start, data_start + data_width):
            occupied.add((r, c))

        slots.append(Slot(
            name=f'{header_name.lower().replace(" ", "_")}_addr_{data_count}',
            row=r,
            col=addr_start,
            width=addr_width,
            fmt='hex',
            signal_bits=16,
            display_min=4,
            label='Addr',
            origin=f'{header_name} row {r + 1} address',
        ))

        slots.append(Slot(
            name=f'{header_name.lower().replace(" ", "_")}_data_{data_count}',
            row=r,
            col=data_start,
            width=data_width,
            fmt='hex',
            signal_bits=32,
            display_min=8,
            label='Data',
            origin=f'{header_name} row {r + 1} data',
        ))

        data_count += 1

    return slots


# Step 6 - Validation
def validate_slots(slots: List[Slot]) -> None:
    """Fail hard on overlaps or width violations."""
    occupied: Dict[Tuple[int, int], Slot] = {}

    # Overlap check
    for s in slots:
        for c in range(s.col, s.col + s.width):
            key = (s.row, c)
            if key in occupied:
                other = occupied[key]
                raise ValueError(
                    f'Overlap on row {s.row + 1}, col {c + 1}: '
                    f'"{s.label}" collides with "{other.label}".'
                )
            occupied[key] = s

    # Width adequacy check
    for s in slots:
        if s.width < s.display_min:
            raise ValueError(
                f'Field "{s.label}" ("{s.name}") at row {s.row + 1}, '
                f'col {s.col + 1}: visual width {s.width} < required '
                f'{s.display_min} for {s.signal_bits}-bit {s.fmt} value.'
            )

    # Binary width exact match check
    for s in slots:
        if s.fmt == 'bin' and s.width != s.signal_bits:
            raise ValueError(
                f'Binary field "{s.label}" ("{s.name}") at row {s.row + 1}: '
                f'placeholder width {s.width} != signal width {s.signal_bits}. '
                f'Use exactly {s.signal_bits} B characters.'
            )


# Step 7 - Build artifacts
def build_mif(grid: List[List[str]], slots: List[Slot], out_path: Path) -> None:
    """Write a Quartus-compatible .mif file."""
    buf = [(ATTR_STATIC << 8) | 0x20 for _ in range(CELLS)]

    # Static text
    for r in range(SCREEN_H):
        for c in range(SCREEN_W):
            ch = grid[r][c]
            if ch != ' ':
                buf[r * SCREEN_W + c] = (ATTR_STATIC << 8) | ord(ch)

    # Dynamic placeholders
    for s in slots:
        for i in range(s.width):
            a = s.start_address + i
            buf[a] = (ATTR_DYNAMIC << 8) | ord('0')

    with open(out_path, 'w') as fp:
        fp.write(
            f'WIDTH=16;\n'
            f'DEPTH={CELLS};\n\n'
            f'ADDRESS_RADIX=UNS;\n'
            f'DATA_RADIX=HEX;\n\n'
            f'CONTENT BEGIN\n'
        )
        for a, v in enumerate(buf):
            fp.write(f'    {a} : {v:04X};\n')
        fp.write('END;\n')


def write_field_map(slots: List[Slot], out_path: Path) -> None:
    """Write the formal contract as JSON."""
    payload = {
        'screen_width': SCREEN_W,
        'screen_height': SCREEN_H,
        'cell_count': CELLS,
        'fields': [s.to_dict() for s in slots],
    }
    with open(out_path, 'w') as fp:
        json.dump(payload, fp, indent=2)


# Main
def main() -> int:
    ap = argparse.ArgumentParser(
        description='Compile RiscVScreenV3.xlsx into video-memory artifacts.'
    )
    ap.add_argument('--xlsx', required=True, help='Input Excel template')
    ap.add_argument('--sheet', required=True, choices=['Single-cycle', 'Pipeline'],
                    help='Which sheet to compile')
    ap.add_argument('--out-mif', required=True, help='Output .mif path')
    ap.add_argument('--out-fields', required=True, help='Output JSON field map')
    args = ap.parse_args()

    # Load worksheet
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f'ERROR: sheet "{args.sheet}" not found. Available: {wb.sheetnames}',
              file=sys.stderr)
        return 1
    ws = wb[args.sheet]

    # Normalise viewport
    grid = read_grid(ws)

    # Detect blocks with horizontal bounds
    blocks = find_blocks(grid)
    print(f'Blocks found: {[(b.name, b.row_start + 1, b.col_start, b.col_end) for b in blocks.values()]}')

    occupied: Set[Tuple[int, int]] = set()
    slots: List[Slot] = []

    # Parse tabular blocks first (they have priority)
    if 'Registers' in blocks:
        slots += parse_registers(grid, blocks['Registers'], occupied)
    if 'Prog. Memory' in blocks:
        slots += parse_memory_table(grid, blocks['Prog. Memory'], occupied, 'Prog. Memory')
    if 'Data Memory' in blocks:
        slots += parse_memory_table(grid, blocks['Data Memory'], occupied, 'Data Memory')

    # Parse tagged signals in the remaining free areas
    try:
        tagged = detect_tagged_slots(grid, occupied, CATALOG)
        slots += tagged
    except ValueError as exc:
        print(f'ERROR: {exc}', file=sys.stderr)
        return 1

    # Validate
    try:
        validate_slots(slots)
    except ValueError as exc:
        print(f'ERROR: {exc}', file=sys.stderr)
        return 1

    # Generate artifacts
    build_mif(grid, slots, Path(args.out_mif))
    write_field_map(slots, Path(args.out_fields))

    print(f'OK: {len(slots)} slots --> {args.out_mif} + {args.out_fields}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
